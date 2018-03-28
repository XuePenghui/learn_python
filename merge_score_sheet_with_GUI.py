# merge two sheet with GUI
from tkinter import *
import tkinter.filedialog
from openpyxl import load_workbook
import xlsxwriter

root = Tk()


def choose_file1():
    filename = tkinter.filedialog.askopenfilename()
#    if filename != '':
    lb1.config(text=filename)
#    else:
#       lb.config(text="您没有选择任何文件")


def choose_file2():
    filename = tkinter.filedialog.askopenfilename()
    lb2.config(text=filename)


def merge_sheet(path1, path2):
    if path1 != '\'\'' and path2 != '\'\'':
        wb1 = load_workbook(path1)
        ws1 = wb1['Sheet1']
        score1 = list(ws1.rows)
        ws1_rows = len(score1)
        ws1_columns = len(score1[1])
        wb2 = load_workbook(path2)
        ws2 = wb2['Sheet1']
        score2 = list(ws2.rows)
        ws2_rows = len(score2)
        ws2_columns = len(score2[1])
        merge_content = []
        for ws1_rows_count in range(ws1_rows):
            for ws2_rows_count in range(ws2_rows):
                if score1[ws1_rows_count][0].value == score2[ws2_rows_count][0].value and score1[ws1_rows_count][1].value == \
                        score2[ws2_rows_count][1].value:
                    content_temp = []
                    for ws1_columns_count in range(ws1_columns):
                        content_temp.append(score1[ws1_rows_count][ws1_columns_count].value)
                    for ws2_columns_count in range(3, ws2_columns):
                        content_temp.append(score2[ws2_rows_count][ws2_columns_count].value)
                    merge_content.append(content_temp)
        workbook = xlsxwriter.Workbook('score_merge.xlsx')
        worksheet = workbook.add_worksheet()
        for row in range(len(merge_content)):
            for col in range(len(merge_content[1])):
                worksheet.write(row, col, merge_content[row][col])
        workbook.close()
        wb1.close()
        wb2.close()
        lb3.config(text="两表格已合并")
    else:
        lb3.config(text="您没有选择正确的文件路径")


lb1 = Label(root, text='')
lb1.pack()
btn1 = Button(root, text="选择文件", command=choose_file1)
btn1.pack()
lb2 = Label(root, text='')
lb2.pack()
btn2 = Button(root, text="选择文件", command=choose_file2)
btn2.pack()
sheet1_temp = lb1['text']
sheet2_temp = lb2['text']
sheet1 = '\'' + sheet1_temp + '\''
sheet2 = '\'' + sheet2_temp + '\''
lb3 = Label(root, text='')
lb3.pack()
btn3 = Button(root, text="合并表格", command=merge_sheet(sheet1, sheet2))
btn3.pack()
root.mainloop()
