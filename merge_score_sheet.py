# merge score1.xlsx and score2.xlsx to score_merge.xlsx
from openpyxl import load_workbook
import xlsxwriter

wb1 = load_workbook('score1.xlsx')
ws1 = wb1['Sheet1']
score1 = list(ws1.rows)
ws1_rows = len(score1)                  # rows of the score1
ws1_columns = len(score1[1])            # columns of the score1

'''
# print content in score1
for rows_count in range(ws1_rows):
    for columns_count in range(ws1_columns):
        print(score1[rows_count][columns_count].value, end='')
        print("\t", end='')
    print("\n")
'''

wb2 = load_workbook('score2.xlsx')
ws2 = wb2['Sheet1']
score2 = list(ws2.rows)
ws2_rows = len(score2)                 # rows of the score2
ws2_columns = len(score2[1])           # columns of the score2

'''
# print content in score2
for rows_count in range(ws2_rows):
    for columns_count in range(ws2_columns):
        print(score2[rows_count][columns_count].value, end='')
        print("\t", end='')
    print("\n")
'''

merge_content = []                      # store the merge result
for ws1_rows_count in range(ws1_rows):
    for ws2_rows_count in range(ws2_rows):
        if score1[ws1_rows_count][0].value == score2[ws2_rows_count][0].value and score1[ws1_rows_count][1].value == score2[ws2_rows_count][1].value and score1[ws1_rows_count][2].value == score2[ws2_rows_count][2].value:
            content_temp = []
            for ws1_columns_count in range(ws1_columns):
                content_temp.append(score1[ws1_rows_count][ws1_columns_count].value)
            for ws2_columns_count in range(3, ws2_columns):
                content_temp.append(score2[ws2_rows_count][ws2_columns_count].value)
            merge_content.append(content_temp)


# write to the new workbook
workbook = xlsxwriter.Workbook('score_merge.xlsx')
worksheet = workbook.add_worksheet()
for row in range(len(merge_content)):
    for col in range(len(merge_content[1])):
        worksheet.write(row, col, merge_content[row][col])
workbook.close()
wb1.close()
wb2.close()
